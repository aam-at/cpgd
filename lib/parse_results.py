import glob
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from config import test_model_thresholds
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .parse_logs import parse_log
from .utils import format_float


def parse_results(model_type,
                  norm,
                  attack,
                  base_dir="results",
                  sort_column=None,
                  group_by=None,
                  extra_params=None):
    if sort_column is None:
        sort_column = f"{norm}_corr"
    if extra_params is None:
        extra_params = []
    if group_by is not None:
        extra_params.append(group_by)
    extra_params.extend(["num_batches", "batch_size"])

    dirs = glob.glob(f"{base_dir}/test_{model_type}/{norm}/{attack}/*")
    df_logs = []
    for load_dir in dirs:
        df_logs.append(
            parse_log(load_dir,
                      exclude=["nll_loss", "conf"],
                      export_test_params=extra_params))
    if len(df_logs) == 0:
        raise ValueError("No logs parsed")
    df = pd.concat(df_logs, ignore_index=True)
    df = df.sort_values(sort_column)
    df["total"] = df["num_batches"] * df["batch_size"]
    df = df.drop(columns=["num_batches", "batch_size"])

    def col_sort(k):
        if k == "name":
            return 1000
        elif k.startswith("acc"):
            return 100
        elif k.startswith(f"{norm}"):
            return 10
        else:
            return 0

    df = df[sorted(df.columns, key=col_sort, reverse=True)]
    for norm in ["l0", "l0p", "l1", "l2", "li"]:
        model_thresholds = np.array(
            test_model_thresholds[model_type][norm if norm != "l0p" else "l0"])
        for col in df.columns:
            if col.startswith(f"acc_{norm}_"):
                threshold = float(col.replace(f"acc_{norm}_", ""))
                if not np.any((model_thresholds - threshold)**2 < 1e-6):
                    df = df.drop(columns=[col])

    for col in df.columns:
        if col.startswith(f"acc_{norm}_"):
            df[col] = df[col] * 100
    return df


def output_excel(df,
                 model_type,
                 norm,
                 attack,
                 base_dir="results",
                 group_by=None):
    df_list = {}
    if group_by is None:
        df_list[attack] = df
    else:
        df_gr = df.groupby(by=group_by)
        for gr_key, gr_df in df_gr:
            gr_df = gr_df.drop(columns=group_by)
            gr_key = format_float(gr_key, 4)
            df_list[f"{attack}-{gr_key}"] = gr_df

    output_file = f"{base_dir}/{model_type}_{norm}.xlsx"
    if Path(output_file).exists():
        writer = pd.ExcelWriter(output_file, mode='a')
        book = load_workbook(output_file)
        if attack in book.sheetnames:
            del book[attack]
        writer.book = book
    else:
        writer = pd.ExcelWriter(output_file, mode='w')

    book = writer.book
    for name, df in df_list.items():
        if name in book.sheetnames:
            del book[name]
        df.to_excel(writer, sheet_name=name)
        # autoresize columns
        ws = writer.sheets[name]
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max(
                        (dims.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value

    def sort_by_title(title):
        if norm == 'li':
            index = {'df': 0, 'bethge': 1, 'daa': 2, 'pgd': 3, 'fab': 4}
        elif norm == 'l2':
            index = {
                'df': 0,
                'cw': 1,
                'ddn': 2,
                'bethge': 3,
                'pgd': 4,
                'fab': 5
            }
        elif norm == 'l1':
            index = {
                'sparsefool': 0,
                'ead': 1,
                'bethge': 2,
                'pgd': 3,
                'fab': 4
            }
        elif norm == 'l0':
            index = {
                'sparsefool': 0,
                'jsma': 1,
                'pixel': 2,
                'bethge': 3,
                'cornersearch': 4,
                'pgd': 5
            }
        else:
            raise ValueError
        for n, v in index.items():
            if n in title:
                return v
        return 100

    book._sheets.sort(key=lambda ws: sort_by_title(ws.title))
    writer.save()

    now = datetime.now()
    today = datetime.today()
    print(f"Wrote {attack}-{norm} on {model_type} at {today}")
